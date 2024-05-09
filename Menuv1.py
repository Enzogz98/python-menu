import os
import tkinter as tk
from tkinter import filedialog, messagebox
from PyPDF2 import PdfReader, PdfWriter
import pandas as pd
import camelot
import aspose.pdf as ap
import re
import shutil
from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import win32com.client as win32
from datetime import datetime
import openpyxl

class PDFApp(tk.Tk):

    
    def __init__(self):
        super().__init__()
        self.title("Convertidor de PDF a Excel")
        self.geometry("400x300")

        self.pdf_dir = tk.StringVar()
        self.generated_excel_files = []  # Lista para almacenar los archivos Excel generados

        self.btn_select_pdf = tk.Button(self, text="Seleccionar Directorio PDF", command=self.select_pdf_directory)
        self.btn_select_pdf.pack(pady=10)

        self.btn_process_files = tk.Button(self, text="Procesar PDFs a Páginas", command=self.process_files, bg='green', fg='white')
        self.btn_process_files.pack(pady=20)
        self.btn_process_files.pack_forget()  # Initially hide this button

        self.btn_convert_to_excel = tk.Button(self, text="Convertir Páginas a Excel", command=self.convert_pages_to_excel, bg='blue', fg='white')
        self.btn_convert_to_excel.pack(pady=10)
        self.btn_convert_to_excel.pack_forget()  # Initially hide this button

        self.btn_unify_and_filter = tk.Button(self, text="Unificar Excel", command=self.unify_and_filter_excel, bg='orange', fg='white')
        self.btn_unify_and_filter.pack(pady=10)
        self.btn_unify_and_filter.pack_forget()

        self.btn_filter_excel = tk.Button(self, text="Filtrar Excel", command=self.filter_excel_files, bg='red', fg='white')
        self.btn_filter_excel.pack(pady=10)
        self.btn_filter_excel.pack_forget()

        self.btn_migrate_to_sheets = tk.Button(self, text="Migrar a Hojas", command=self.migrate_to_sheets, bg='purple', fg='white')
        self.btn_migrate_to_sheets.pack(pady=10)
        self.btn_migrate_to_sheets.pack_forget() 
        

    def select_pdf_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.pdf_dir.set(directory)
            self.btn_process_files.pack()  # Show the process button

    def process_files(self):
        if not self.pdf_dir.get():
            messagebox.showerror("Error", "Por favor, seleccione un directorio antes de procesar.")
            return
        
        pdf_directory = self.pdf_dir.get()
        self.pdf_pages_directory = os.path.join(pdf_directory, "pdfPages")
        
        os.makedirs(self.pdf_pages_directory, exist_ok=True)
        self.split_all_pdfs(pdf_directory, self.pdf_pages_directory)

        messagebox.showinfo("Éxito", "Todos los PDFs han sido procesados a páginas exitosamente!")
        self.btn_process_files.pack_forget()  # Hide the process button
        self.btn_convert_to_excel.pack()  # Show the convert to Excel button

    def migrate_to_sheets(self):
            if self.generated_excel_files:
                combined_dir = os.path.join(os.path.dirname(self.generated_excel_files[0]), "excelCombined")
                output_folder_path = os.path.join(combined_dir, 'excelFiltered')

                save_file_path = filedialog.asksaveasfilename(title="Guardar Archivo Maestro", defaultextension=".xlsm", filetypes=[("Excel Files", "*.xlsm")])
                if not save_file_path:
                    messagebox.showerror("Error", "Por favor, seleccione una ubicación y nombre para el archivo maestro.")
                    return

                base_path = filedialog.askdirectory(title="Seleccionar Carpeta de Destino")
                if not base_path:
                    messagebox.showerror("Error", "Por favor, seleccione una carpeta de destino.")
                    return

                self.update_master_sheet("C:/Python312/CONVERSOR DE PLANILLA SAS v.5.xlsm", output_folder_path, save_file_path, base_path)
            else:
                messagebox.showerror("Error", "No se han generado archivos Excel para migrar.")

    def adapt_formula(self, formula, new_row):
        cell_ref_pattern = r'(\$?[A-Za-z]+)(\$?\d+)'

        def replace_cell_reference(match):
            col_letter, row_number = match.group(1), match.group(2)
            if '$' in col_letter or '$' in row_number:
                return f"{col_letter}{row_number}"
            else:
                adjusted_row_number = new_row if int(row_number.lstrip('$')) == 2 else int(row_number) + new_row - 2
                return f"{col_letter}{adjusted_row_number}"

        formula = re.sub(r'([A-Z]+)(\d+)', lambda x: replace_cell_reference(x), formula)
        return formula

    def run_macro(self, excel_path, macro_name):
        excel = win32.Dispatch("Excel.Application")
        workbook = excel.Workbooks.Open(excel_path)
        excel.Visible = True  
        excel.Application.Run(f"{macro_name}")
        workbook.Save()
        workbook.Close()

    def copy_sas_sheet_only(self, source_path, base_path):
        wb = openpyxl.load_workbook(source_path)
        ws = wb['SAS']
        date_cell_value = ws['C2'].value
        date_style = ws['A2'].number_format  
        
        if isinstance(date_cell_value, datetime):
            date_str = date_cell_value.strftime('%Y%m%d')
        else:
            date_str = datetime.now().strftime('%Y%m%d')  
        new_filename = f"SAS_{date_str}.xlsx"
        new_file_path = os.path.join(base_path, new_filename)
        
        new_wb = Workbook()
        new_ws = new_wb.active
        for row in ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = copy(cell.alignment)
        
        for row in new_ws.iter_rows(min_row=2, max_col=1, max_row=new_ws.max_row):
            for cell in row:
                cell.number_format = date_style
        
        new_wb.save(new_file_path)
        new_wb.close()
        return new_file_path



    

    def update_master_sheet(self, master_path, folder_path, save_path,base_path):
        if not os.path.exists(save_path):
            shutil.copy(master_path, save_path)
            print("Copia creada en:", save_path)
        else:
            print("Copia ya existe en:", save_path)
        wb_maestro = openpyxl.load_workbook(save_path, keep_vba=True)



        name_to_sheet = {
            'VISA DEBIT': 'Visa debito',
            'VISA': 'Visa',
            'MASTERCARD DEBIT': 'Mastercard debito',
            'MAESTRO': 'MAESTRO',
            'MASTERCARD': 'Mastercard',
            'CABAL': 'CABAL',
            'CABAL DEBIT': 'CABAL',
            'AMEX': 'AMEX FISERV',
            'ARGENCARD': 'ARGENCARD'
        }

        sheets_with_data = {}
        sheets_with_data_bool = {}  

        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            if os.path.isfile(file_path):
                for key, value in name_to_sheet.items():
                    if key in filename:
                        sheet_name = value
                        break
                else:
                    continue

                wb_current = openpyxl.load_workbook(file_path, data_only=False)
                sheet_current = wb_current.active
                sheet_maestro = wb_maestro[sheet_name]

                if sheet_maestro.max_row > 2:
                    sheet_maestro.delete_rows(3, sheet_maestro.max_row - 1)

                formulas = {}
                for col_idx in range(1, 47):  
                    cell = sheet_maestro.cell(row=2, column=col_idx)
                    if cell.data_type == 'f':
                        formulas[col_idx] = cell.value

                data_added = False
                for row_idx, row in enumerate(sheet_current.iter_rows(min_row=3, max_col=47), start=3):
                    for col_idx, cell in enumerate(row, start=1):
                        new_cell = sheet_maestro.cell(row=row_idx, column=col_idx)
                        if col_idx in formulas:
                            adapted_formula = self.adapt_formula(formulas[col_idx], row_idx)
                            new_cell.value = '=' + adapted_formula
                        else:
                            new_cell.value = cell.value
                        if cell.value:
                            data_added = True

                        if row_idx > 1 and col_idx > 1:
                            prev_cell = sheet_maestro.cell(row=row_idx - 1, column=col_idx)
                            new_cell.number_format = prev_cell.number_format
                            new_cell.font = copy(prev_cell.font)
                            new_cell.alignment = copy(prev_cell.alignment)
                            new_cell.border = copy(prev_cell.border)
                            new_cell.fill = copy(prev_cell.fill)

                sheets_with_data[sheet_name] = data_added
                sheets_with_data_bool[sheet_name] = data_added  
                
                last_row_with_data = sheet_maestro.max_row
                table_name = self.get_table_name_for_sheet(sheet_name)
                if table_name:
                    table = sheet_maestro.tables.get(table_name)
                    if table:
                        table_range = table.ref.split(':')
                        table_ref_start = table_range[0]
                        table_ref_end = table_range[1]
                        updated_table_range = f"{table_ref_start}:{table_ref_end.split('$')[0]}{last_row_with_data}"
                        if re.match(r'^[$]?([A-Za-z]{1,3})[$]?(\d+)(:[$]?([A-Za-z]{1,3})[$]?(\d+)?)?$|^[A-Za-z]{1,3}:[A-Za-z]{1,3}$', updated_table_range):
                            table.ref = updated_table_range
                        else:
                            print(f"La referencia de la tabla '{table_name}' no es válida: '{updated_table_range}'")


        wb_maestro.save(save_path)
        for sheet, has_data in sheets_with_data_bool.items():
            if has_data:
                macro_name = f"{sheet.replace(' ', '_')}"
                self.run_macro(save_path, macro_name)
                print(f"Macro {macro_name} ejecutada para {sheet}")


        self.run_macro(save_path, "ConvertirFechaCorta")
        print("Macro 'ConvertirFechaCorta' ejecutada para la hoja 'SAS'")

        final_path = self.copy_sas_sheet_only(save_path, base_path)
        final_wb = openpyxl.load_workbook(final_path)
        final_sheet = final_wb.active
        self.auto_adjust_column_width(final_sheet)
        final_wb.save(final_path)

        return sheets_with_data, sheets_with_data_bool

    def get_table_name_for_sheet(self, sheet_name):
        table_names = {
            'Visa debito': 'Tabla14',
            'Visa': 'Tabla1',
            'Mastercard debito': 'Tabla145',
            'MAESTRO': 'Tabla1456',
            'Mastercard': 'Tabla13',
            'CABAL': 'Tabla7',
            'AMEX FISERV': 'Tabla19',
            'ARGENCARD': 'Tabla137'
        }
        return table_names.get(sheet_name, None)

    def auto_adjust_column_width(self, sheet):
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                         max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    def unify_and_filter_excel(self):
        if self.generated_excel_files:
            combined_dir = os.path.join(os.path.dirname(self.generated_excel_files[0]), "excelCombined")
            os.makedirs(combined_dir, exist_ok=True)

            self.combine_and_save_by_tarjeta(self.generated_excel_files, combined_dir)

            

            messagebox.showinfo("Éxito", "Archivos Excel unificados exitosamente!")
            self.btn_unify_and_filter.pack_forget()  
            self.btn_filter_excel.pack()
        else:
            messagebox.showerror("Error", "No se han generado archivos Excel para unificar y filtrar.")

    def combine_and_save_by_tarjeta(self, excel_files, output_dir):
        nombres_tarjetas = ["VISA DEBIT", "VISA", "MASTERCARD DEBIT", "MASTERCARD", "ARGENCARD", "CABAL DEBIT", "CABAL", "AMEX", "MAESTRO"]
        dfs_tarjetas = {nombre: pd.DataFrame() for nombre in nombres_tarjetas}

        for nombre_tarjeta in nombres_tarjetas:
            archivos = [f for f in excel_files if self.nombre_en_archivo(nombre_tarjeta, f)]
            print(f"Archivos encontrados para {nombre_tarjeta}: {archivos}")
            for archivo in archivos:
                ruta_completa = os.path.join(output_dir, archivo)
                df_temp = pd.read_excel(ruta_completa, header=None)
                rows = df_temp.values.tolist()
                for row in rows:
                    while row[0]=="":
                        row=row[1:]+[""]
                        
                        if not any(row):
                            break

                dfs_tarjetas[nombre_tarjeta] = pd.concat([dfs_tarjetas[nombre_tarjeta], df_temp])

        for nombre_tarjeta, df in dfs_tarjetas.items():
            if not df.empty:
                output_path = os.path.join(output_dir, f"{nombre_tarjeta}_combined.xlsx")
                df.to_excel(output_path, index=False)
                print(f'Archivo guardado: {output_path}')

    def nombre_en_archivo(self, nombre_tarjeta, nombre_archivo):
        if "DEBIT" in nombre_tarjeta:
            pattern = rf'\b{re.escape(nombre_tarjeta)}(?:_|\b)'
        else:
            debit_free_pattern = rf'\b{re.escape(nombre_tarjeta)}(?:_|\b)(?!\sDEBIT)'
            pattern = debit_free_pattern if nombre_tarjeta in ["VISA", "MASTERCARD"] else rf'\b{re.escape(nombre_tarjeta)}(?:_|\b)'
        result = re.search(pattern, nombre_archivo) is not None
        print(f"Checking {nombre_archivo} for {nombre_tarjeta} using pattern {pattern}: {result}")
        return result

    def split_all_pdfs(self, pdf_directory, pdf_pages_directory):
        for filename in os.listdir(pdf_directory):
            if filename.lower().endswith('.pdf'):
                pdf_path = os.path.join(pdf_directory, filename)
                self.split_pdf_pages(pdf_path, pdf_pages_directory)

    def split_pdf_pages(self, pdf_path, output_directory):
        with open(pdf_path, 'rb') as file:
            pdf_reader = PdfReader(file)
            num_pages = len(pdf_reader.pages)
            for page_num in range(num_pages):
                page = pdf_reader.pages[page_num]
                output_page_path = os.path.join(output_directory, f"{os.path.splitext(os.path.basename(pdf_path))[0]}_página_{page_num + 1}.pdf")
                pdf_writer = PdfWriter()
                pdf_writer.add_page(page)
                with open(output_page_path, 'wb') as output_file:
                    pdf_writer.write(output_file)

    def convert_pages_to_excel(self):
        excel_directory = os.path.join(os.path.dirname(self.pdf_pages_directory), "pagesExcel")
        os.makedirs(excel_directory, exist_ok=True)
        self.process_pdfs_from_directory(self.pdf_pages_directory, excel_directory)
        messagebox.showinfo("Éxito", "Todas las páginas han sido convertidas a Excel exitosamente!")
        self.btn_convert_to_excel.pack_forget()
        self.generated_excel_files = [os.path.join(excel_directory, f) for f in os.listdir(excel_directory) if f.endswith('.xlsx')]
        self.btn_unify_and_filter.pack()

    def process_pdfs_from_directory(self, pdf_directory, excel_directory):
        for filename in os.listdir(pdf_directory):
            if filename.lower().endswith('.pdf'):
                pdf_path = os.path.join(pdf_directory, filename)
                excel_filename = filename[:-4] + '.xlsx'
                excel_path = os.path.join(excel_directory, excel_filename)
                self.convert_pdf_to_excel(pdf_path, excel_path)

    def convert_pdf_to_excel(self, pdf_path, excel_path):
        tables = camelot.read_pdf(pdf_path, flavor='lattice', pages='all')
        all_data_frames = [] 
        if tables.n > 0:
            for table in tables:
                df = table.df
                df.columns = df.iloc[0]  
                df = df[1:].reset_index(drop=True)
                df.replace('', pd.NA, inplace=True)
                if "MAESTRO" in pdf_path.upper():
                    df.iloc[:, 2] = df.iloc[:, 2].replace(r'\n', ' ', regex=True) 
                if not df.empty:
                    all_data_frames.append(df)
            final_df = pd.concat(all_data_frames, ignore_index=True)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                final_df.to_excel(writer, sheet_name='Todas las Tablas', index=False)
            print(f'{pdf_path} convertido a {excel_path} con Camelot en una sola hoja')
        else:
            self.convert_pdf_to_excel_with_aspose(pdf_path, excel_path)

    def convert_pdf_to_excel_with_aspose(self, pdf_path, excel_path):
        document = ap.Document(pdf_path)
        excel_save_options = ap.ExcelSaveOptions()
        excel_save_options.format = ap.ExcelSaveOptions.ExcelFormat.XLSX
        document.save(excel_path, excel_save_options)
        print(f'{pdf_path} convertido a Excel con Aspose en {excel_path}')

        if "MAESTRO" in pdf_path.upper():
            self.adjust_columns_excel(excel_path)

    def adjust_columns_excel(self, excel_path):
        try:
            data = pd.read_excel(excel_path, header=0)
            if data.shape[1] >= 8:
                data.iloc[:, 2] =data.iloc[:,2].astype(str) + ' ' + data.iloc[:, 3].astype(str) + ' ' + data.iloc[:, 4].astype(str)
                data.iloc[:, 3] = data.iloc[:, 5]
                data.iloc[:, 4] = data.iloc[:, 6]
                data.iloc[:, 5] = data.iloc[:, 7]
                data.drop(data.columns[[6, 7]], axis=1, inplace=True)
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    data.to_excel(writer, index=False)
                print(f'Ajuste de columnas realizado en {excel_path}')
        except Exception as e:
            print(f"Error al ajustar las columnas en {excel_path}: {e}")
            
 

    def convert_to_float(self,x):
        if pd.notna(x) and isinstance(x, str):
            try:
                return float(x.replace('.', '').replace(',', '.'))
            except ValueError:
                return x  
        return x

    def separate_lote_cupon(self, data):
        split_content = data.iloc[:, 2].str.split(expand=True)
        data.insert(3, 'lote', split_content[1])
        data.insert(4, 'cupon', split_content[2]) if split_content.shape[1] > 1 else None
        data.iloc[:, 2] = split_content[0] if split_content.shape[1] > 2 else None
        return data

    def clean_empty_cells(self, data, start_col, sheet_name):
        for col in range(start_col, len(data.columns)):
            data.iloc[:, col] = data.iloc[:, col].apply(lambda x: x if pd.notna(x) and str(x).strip() != '' else None)
        if 'MAESTRO' in sheet_name:
            data['Term/Lote/Cupon'] = data['Term/Lote/Cupon'].apply(lambda x: x if pd.notna(x) and str(x).strip() != '' else "Default Value")
        return data

    def apply_numeric_conversion(self, data):
        for col in data.columns:
            data[col] = data[col].apply(self.convert_to_float)
        return data

    def filter_excel(self, file_path):

        fixed_headers = [
    "Trx", "Fecha Pres Fecha", "Term/Lote/Cupon", "Tarj", "Plan Cuota", "T F", 
    "T.N.A. %", "Ventas con/Dto.", "Ventas sin/Dto.", "Dto. Arancel", 
    "Dto. Financ.", "Cod. Rechazo Mot. contrap."
]
        try:
            data = pd.read_excel(file_path, header=1)
            
            if data.shape[1] > 12:
                data = data.iloc[:, :-1]
            
            current_headers = fixed_headers[:min(len(fixed_headers), len(data.columns))]
            data.columns = current_headers
            
            if 'MAESTRO' in file_path:
                data = self.separate_lote_cupon(data)
                data = self.clean_empty_cells(data, 2, 'MAESTRO')
            else:
                data = self.clean_empty_cells(data, 3, 'Other')
            
            data = self.apply_numeric_conversion(data)
            
            if data.iloc[:, 0].isin(['Plan cuota', 'Venta ctdo']).any():
                filtered_data = data[data.iloc[:, 0].isin(['Plan cuota', 'Venta ctdo'])]
                return filtered_data
            else:
                print(f"No se encontraron entradas válidas en la primera columna del archivo: {file_path}")
                return None
        except Exception as e:
            print(f"Error procesando el archivo {file_path}: {e}")
            return None

    def filter_excel_files(self):
        combined_dir=os.path.join(os.path.dirname(self.generated_excel_files[0]), "excelCombined")
        if os.path.exists(combined_dir):
            output_folder_path = os.path.join(combined_dir, 'excelFiltered')
            os.makedirs(output_folder_path, exist_ok=True)
            self.filter_combined_excel_files(combined_dir, output_folder_path)
            messagebox.showinfo("Éxito", "Archivos Excel combinados filtrados exitosamente!")
            self.btn_filter_excel.pack_forget()
            self.btn_migrate_to_sheets.pack()
        else:
            messagebox.showerror("Error", "No se encontró la carpeta de archivos combinados.")
    def filter_combined_excel_files(self, combined_dir, output_folder):
        combined_files = [f for f in os.listdir(combined_dir) if f.endswith('.xlsx')]
        for combined_file in combined_files:
            combined_file_path = os.path.join(combined_dir, combined_file)
            filtered_data = self.filter_excel(combined_file_path)
            if filtered_data is not None:
                output_path = os.path.join(output_folder, f'filtered_{combined_file}')
                filtered_data.to_excel(output_path, index=False)
                print(f'Archivo combinado filtrado guardado en: {output_path}')



        
        
if __name__ == "__main__":
    app = PDFApp()
    app.mainloop()

