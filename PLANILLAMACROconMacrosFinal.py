import openpyxl
from openpyxl import Workbook

import os
import re
import shutil
from copy import copy
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
import win32com.client as win32
from datetime import datetime



def adapt_formula(formula, new_row):
    cell_ref_pattern = r'(\$?[A-Za-z]+)(\$?\d+)'

    def replace_cell_reference(match):
        col_letter, row_number = match.group(1), match.group(2)
        if '$' in col_letter or '$' in row_number:
            return f"{col_letter}{row_number}"
        else:
            adjusted_row_number = new_row if int(row_number.lstrip('$')) == 2 else int(row_number) + new_row - 2
            return f"{col_letter}{adjusted_row_number}"

    formula = re.sub(r'([A-Z]+)(\d+)', lambda x: replace_cell_reference(x), formula)
    return formula

def run_macro(excel_path, macro_name):
    excel = win32.Dispatch("Excel.Application")
    workbook = excel.Workbooks.Open(excel_path)
    excel.Visible = True  
    excel.Application.Run(f"{macro_name}")
    workbook.Save()
    workbook.Close()

def copy_sas_sheet_only(source_path, base_path):
    wb = openpyxl.load_workbook(source_path)
    ws = wb['SAS']
    date_cell_value = ws['C2'].value
    date_style = ws['A2'].number_format  
    
    if isinstance(date_cell_value, datetime):
        date_str = date_cell_value.strftime('%Y%m%d')
    else:
        date_str = datetime.now().strftime('%Y%m%d')  
    new_filename = f"SAS_{date_str}.xlsx"
    new_file_path = os.path.join(base_path, new_filename)
    
    new_wb = Workbook()
    new_ws = new_wb.active
    for row in ws.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.alignment = copy(cell.alignment)
    
    for row in new_ws.iter_rows(min_row=2, max_col=1, max_row=new_ws.max_row):
        for cell in row:
            cell.number_format = date_style
    
    new_wb.save(new_file_path)
    new_wb.close()
    return new_file_path



def auto_adjust_column_width(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column  
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

def update_master_sheet(master_path, folder_path, save_path,base_path):
    if not os.path.exists(save_path):
        shutil.copy(master_path, save_path)
        print("Copia creada en:", save_path)
    else:
        print("Copia ya existe en:", save_path)
    wb_maestro = openpyxl.load_workbook(save_path, keep_vba=True)



    name_to_sheet = {
        'VISA DEBIT': 'Visa debito',
        'VISA': 'Visa',
        'MASTERCARD DEBIT': 'Mastercard debito',
        'MAESTRO': 'MAESTRO',
        'MASTERCARD': 'Mastercard',
        'CABAL': 'CABAL',
        'CABAL DEBIT': 'CABAL',
        'AMEX': 'AMEX FISERV',
        'ARGENCARD': 'ARGENCARD'
    }

    sheets_with_data = {}
    sheets_with_data_bool = {}  

    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if os.path.isfile(file_path):
            for key, value in name_to_sheet.items():
                if key in filename:
                    sheet_name = value
                    break
            else:
                continue

            wb_current = openpyxl.load_workbook(file_path, data_only=False)
            sheet_current = wb_current.active
            sheet_maestro = wb_maestro[sheet_name]

            if sheet_maestro.max_row > 2:
                sheet_maestro.delete_rows(3, sheet_maestro.max_row - 1)

            formulas = {}
            for col_idx in range(1, 47):  
                cell = sheet_maestro.cell(row=2, column=col_idx)
                if cell.data_type == 'f':
                    formulas[col_idx] = cell.value

            data_added = False
            for row_idx, row in enumerate(sheet_current.iter_rows(min_row=3, max_col=47), start=3):
                for col_idx, cell in enumerate(row, start=1):
                    new_cell = sheet_maestro.cell(row=row_idx, column=col_idx)
                    if col_idx in formulas:
                        adapted_formula = adapt_formula(formulas[col_idx], row_idx)
                        new_cell.value = '=' + adapted_formula
                    else:
                        new_cell.value = cell.value
                    if cell.value:
                        data_added = True

                    if row_idx > 1 and col_idx > 1:
                        prev_cell = sheet_maestro.cell(row=row_idx - 1, column=col_idx)
                        new_cell.number_format = prev_cell.number_format
                        new_cell.font = copy(prev_cell.font)
                        new_cell.alignment = copy(prev_cell.alignment)
                        new_cell.border = copy(prev_cell.border)
                        new_cell.fill = copy(prev_cell.fill)

            sheets_with_data[sheet_name] = data_added
            sheets_with_data_bool[sheet_name] = data_added  
            
            last_row_with_data = sheet_maestro.max_row
            table_name = get_table_name_for_sheet(sheet_name)
            if table_name:
                table = sheet_maestro.tables.get(table_name)
                if table:
                    table_range = table.ref.split(':')
                    table_ref_start = table_range[0]
                    table_ref_end = table_range[1]
                    updated_table_range = f"{table_ref_start}:{table_ref_end.split('$')[0]}{last_row_with_data}"
                    if re.match(r'^[$]?([A-Za-z]{1,3})[$]?(\d+)(:[$]?([A-Za-z]{1,3})[$]?(\d+)?)?$|^[A-Za-z]{1,3}:[A-Za-z]{1,3}$', updated_table_range):
                        table.ref = updated_table_range
                    else:
                        print(f"La referencia de la tabla '{table_name}' no es válida: '{updated_table_range}'")


    wb_maestro.save(save_path)
    for sheet, has_data in sheets_with_data_bool.items():
        if has_data:
            macro_name = f"{sheet.replace(' ', '_')}"
            run_macro(save_path, macro_name)
            print(f"Macro {macro_name} ejecutada para {sheet}")


    run_macro(save_path, "ConvertirFechaCorta")
    print("Macro 'ConvertirFechaCorta' ejecutada para la hoja 'SAS'")

    final_path = copy_sas_sheet_only(save_path, base_path)
    final_wb = openpyxl.load_workbook(final_path)
    final_sheet = final_wb.active
    auto_adjust_column_width(final_sheet)
    final_wb.save(final_path)

    return sheets_with_data, sheets_with_data_bool

def get_table_name_for_sheet(sheet_name):
    table_names = {
        'Visa debito': 'Tabla14',
        'Visa': 'Tabla1',
        'Mastercard debito': 'Tabla145',
        'MAESTRO': 'Tabla1456',
        'Mastercard': 'Tabla13',
        'CABAL': 'Tabla7',
        'AMEX FISERV': 'Tabla19',
        'ARGENCARD': 'Tabla137'
    }
    return table_names.get(sheet_name, None)

master_file_path = 'C:/Python312/CONVERSOR DE PLANILLA SAS v.5.xlsm'
source_folder_path = 'C:/Python312/excelConverted/excelCrudo'
save_file_path = 'C:/Python312/Copia_CONVERSOR_DE_PLANILLA_SAS_v.5.xlsm'
base_path = 'C:/Python312/'


sheets_with_data, sheets_with_data_bool = update_master_sheet(master_file_path, source_folder_path, save_file_path, base_path)


print("Hoja(s) con datos ingresados:")
for sheet_name, data_added in sheets_with_data.items():
    if data_added:
        print(sheet_name)

print("Variables booleanas para hojas con datos ingresados:")
for sheet_name, has_data in sheets_with_data_bool.items():
    print(f"{sheet_name}: {has_data}")

